"""Document parser tool — PDF, Word, and Excel parsing."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Union

from src.tools.base_tool import BaseTool, ToolMetadata

logger = logging.getLogger(__name__)


class DocumentParserTool(BaseTool):
    """Parses documents (PDF, Word, Excel) and extracts text and structure."""

    @property
    def metadata(self) -> ToolMetadata:
        return ToolMetadata(
            name="document_parser",
            description="Parse PDF, Word, and Excel documents and extract text",
            parameters={
                "required": ["file_path"],
                "optional": ["output_format"],
                "properties": {
                    "file_path": {"type": "string", "description": "Path to the document"},
                    "output_format": {
                        "type": "string",
                        "enum": ["text", "structured"],
                        "default": "text",
                    },
                },
            },
            return_type="Dict",
            category="information_retrieval",
        )

    def execute(self, **kwargs: Any) -> Dict[str, Any]:
        if not self.validate_parameters(**kwargs):
            return {"error": "Missing required parameters", "content": None}

        file_path = Path(kwargs["file_path"])
        output_format = kwargs.get("output_format", "text")

        if not file_path.exists():
            return {"error": f"File not found: {file_path}", "content": None}

        suffix = file_path.suffix.lower()
        try:
            if suffix == ".pdf":
                return self._parse_pdf(file_path, output_format)
            if suffix in (".doc", ".docx"):
                return self._parse_word(file_path, output_format)
            if suffix in (".xls", ".xlsx"):
                return self._parse_excel(file_path, output_format)
            # Fallback: plain text
            return self._parse_text(file_path)
        except Exception as exc:  # noqa: BLE001
            logger.error("Document parsing failed for '%s': %s", file_path, exc)
            return {"error": str(exc), "content": None}

    # ------------------------------------------------------------------
    # Parsers
    # ------------------------------------------------------------------

    def _parse_pdf(self, path: Path, fmt: str) -> Dict[str, Any]:
        try:
            import pypdf  # type: ignore[import]

            reader = pypdf.PdfReader(str(path))
            pages = [page.extract_text() or "" for page in reader.pages]
            content: Union[str, List[str]] = pages if fmt == "structured" else "\n".join(pages)
            return {"file": str(path), "type": "pdf", "pages": len(pages), "content": content}
        except ImportError:
            logger.warning("pypdf not installed; returning mock PDF content")
            return {"file": str(path), "type": "pdf", "content": "[PDF content placeholder]"}

    def _parse_word(self, path: Path, fmt: str) -> Dict[str, Any]:
        try:
            import docx  # type: ignore[import]

            doc = docx.Document(str(path))
            paragraphs = [p.text for p in doc.paragraphs]
            content: Union[str, List[str]] = paragraphs if fmt == "structured" else "\n".join(paragraphs)
            return {"file": str(path), "type": "docx", "paragraphs": len(paragraphs), "content": content}
        except ImportError:
            logger.warning("python-docx not installed; returning mock Word content")
            return {"file": str(path), "type": "docx", "content": "[Word content placeholder]"}

    def _parse_excel(self, path: Path, fmt: str) -> Dict[str, Any]:
        try:
            import openpyxl  # type: ignore[import]

            wb = openpyxl.load_workbook(str(path), data_only=True)
            sheets: Dict[str, List[List[Any]]] = {}
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sheets[sheet] = [[cell.value for cell in row] for row in ws.iter_rows()]
            if fmt == "structured":
                return {"file": str(path), "type": "xlsx", "sheets": sheets}
            rows = [
                "\t".join(str(c) for c in row)
                for rows in sheets.values()
                for row in rows
            ]
            return {"file": str(path), "type": "xlsx", "content": "\n".join(rows)}
        except ImportError:
            logger.warning("openpyxl not installed; returning mock Excel content")
            return {"file": str(path), "type": "xlsx", "content": "[Excel content placeholder]"}

    def _parse_text(self, path: Path) -> Dict[str, Any]:
        content = path.read_text(encoding="utf-8", errors="replace")
        return {"file": str(path), "type": "text", "content": content}
